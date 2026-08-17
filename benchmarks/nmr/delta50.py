"""DELTA50 as ground truth: 50 small molecules, assigned 13C in CDCl3.

SOURCE [source:kwon2023]. Kwon, Y. et al. "DELTA50: A Highly Accurate Database of
Experimental 1H and 13C NMR Chemical Shifts Applied to DFT Benchmarking",
Molecules 2023, 28, 2449, doi:10.3390/molecules28062449 (CC BY 4.0).
Shifts read from `DELTA50_benchmark.xlsx` in the paper's supplementary
archive: 600 MHz, CDCl3, <=10 mM, TMS internal reference, ambiguities
resolved by gCOSY/gHSQC/gHMBC.

WHY THIS SET AND NOT nmrshiftdb2. nmrshiftdb2 IS the lookup's index, so
scoring the lookup against it would be circular. DELTA50 is independent of
both methods under test, measured in the solvent everything else here uses,
and was built specifically so that DFT could be benchmarked against it.

THE HARD PART IS THE MAPPING, not the numbers. DELTA50 labels carbons by
its own numbering ("1", "2,3", "1-3"), which is not RDKit's. Two guards:

  1. The row count per compound must equal the number of carbon symmetry
     classes RDKit finds. This catches a wrong SMILES immediately -- it
     passed for 47 of 50 and the three failures were all real chemistry
     (below), not typos.
  2. The assignment itself comes from matching OUR computed shieldings to
     DELTA50's OWN computed shieldings, both on the same physical scale.
     Deliberately NOT from matching against the lookup: using one of the
     methods under test to decide what the truth is would bias the
     comparison toward it.
"""

from __future__ import annotations

import json
import zipfile
from dataclasses import dataclass
from pathlib import Path

from rdkit import Chem

HERE = Path(__file__).resolve().parent
DATA = HERE / "delta50_13c.json"

#: Written out rather than resolved from the names at run time: several are
#: abbreviations or house names no name-to-structure parser handles, and one
#: is actively misleading -- DELTA50's "1,2-Pyrazine" is pyridazine.
#: Every entry is checked by the symmetry-class count above.
SMILES: dict[str, str] = {
    "1,2-Pyrazine": "c1ccnnc1",
    "1,4-Benzoquinone": "O=C1C=CC(=O)C=C1",
    "1,4-Pyrazine": "c1cnccn1",
    "2,5-Dihydrofuran": "C1C=CCO1",
    "2-Butyne": "CC#CC",
    "2-Cyanopropane": "CC(C)C#N",
    "3-Butyn-2-one": "CC(=O)C#C",
    "Acetaldehyde": "CC=O",
    "Acetone": "CC(C)=O",
    "Acetonitrile": "CC#N",
    "Anisole": "COc1ccccc1",
    "Benzene": "c1ccccc1",
    "Butyrolactone": "O=C1CCCO1",
    "Cyclobutanone": "O=C1CCC1",
    "Cyclohex-2-en-1-one": "O=C1CCCC=C1",
    "Cyclohexane": "C1CCCCC1",
    "Cyclohexanone": "O=C1CCCCC1",
    "Cyclopentane": "C1CCCC1",
    "Cyclopentanone": "O=C1CCCC1",
    "Cyclopentenone": "O=C1CCC=C1",
    "Cyclopropane": "C1CC1",
    "Fluorobenzene": "Fc1ccccc1",
    "Furan": "c1ccoc1",
    "Isobutylene": "CC(C)=C",
    "Isoxazole": "c1ccno1",
    "MTBE": "COC(C)(C)C",
    "Maleic anhydride": "O=C1OC(=O)C=C1",
    "Methyl acetate": "COC(C)=O",
    "N-Methylpiperidine": "CN1CCCCC1",
    "N-Methylpyrrole": "Cn1cccc1",
    "N-Methylpyrrolidine": "CN1CCCC1",
    "Nitrobenzene": "O=[N+]([O-])c1ccccc1",
    "Nitroethane": "CC[N+](=O)[O-]",
    "Nitromethane": "C[N+](=O)[O-]",
    "Norbornadiene": "C1=CC2C=CC1C2",
    "Oxetane": "C1COC1",
    "Oxirane": "C1CO1",
    "Pivalonitrile": "CC(C)(C)C#N",
    "Propionitrile": "CCC#N",
    "Pyridine": "c1ccncc1",
    "Pyrimidine": "c1cncnc1",
    "THF": "C1CCOC1",
    "THP": "C1CCOCC1",
    "Toluene": "Cc1ccccc1",
    "t-Butyl nitrate": "CC(C)(C)O[N+](=O)[O-]",
    "t-Butylacetylene": "CC(C)(C)C#C",
    "t-Butylethylene": "CC(C)(C)C=C",
}

#: Excluded, with the reason, rather than silently absent. All three fail
#: the symmetry-class check for the SAME chemical reason: carbons that the
#: molecular graph makes equivalent but that experiment resolves, because
#: something restricts their interconversion. Nothing in this project's
#: graph-based machinery can tell such a pair apart, so any assignment
#: would be a coin flip -- and a wrong one is worse than a missing one.
EXCLUDED: dict[str, str] = {
    "DMF": (
        "3 measured carbons, 2 graph classes: amide C-N rotation is "
        "restricted, so the two N-methyls are cis and trans to the "
        "carbonyl and resolve separately."
    ),
    "DMAc": "As DMF -- restricted amide rotation splits the two N-methyls.",
    "2-Methyl-2-butene": (
        "5 measured carbons, 4 graph classes: the two methyls on the sp2 "
        "carbon sit E and Z across the double bond."
    ),
}


@dataclass(frozen=True)
class Delta50Compound:
    name: str
    smiles: str
    #: One entry per carbon symmetry class, in DELTA50's row order:
    #: (their label, experimental ppm, their computed isotropic shielding).
    entries: list[tuple[str, float, float]]


def extract(zip_path: Path, out: Path = DATA) -> dict:
    """Pull the 13C table out of the supplementary workbook, once."""
    import io

    import openpyxl

    # Read the whole workbook into memory first: openpyxl's read-only mode
    # pulls sheet XML lazily, so a live zip handle is closed out from under
    # it before the first row is touched.
    with zipfile.ZipFile(zip_path) as archive:
        payload = archive.read("DELTA50_benchmark.xlsx")
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook["13C Solvent"]
    compounds: dict[str, list] = {}
    current = None
    # Row 38 is the header; the per-carbon rows run from 39 to the end.
    for row in sheet.iter_rows(min_row=39, max_row=sheet.max_row, max_col=4, values_only=True):
        name, label, experimental, shielding = row[0], row[1], row[2], row[3]
        if name:
            current = str(name).strip()
            compounds.setdefault(current, [])
        if current is None or experimental is None or label is None:
            continue
        compounds[current].append([str(label).strip(), float(experimental), float(shielding)])
    compounds = {k: v for k, v in compounds.items() if v}
    out.write_text(json.dumps(compounds, indent=1, sort_keys=True), encoding="utf-8")
    return compounds


def carbon_classes(mol: Chem.Mol) -> list[list[int]]:
    """Carbon symmetry classes, each a list of atom indices.

    `breakTies=False` is what makes this a symmetry perception rather than
    a canonical ordering -- the same call `chem/nmr_signals.py` already
    uses to group equivalent protons.
    """
    ranks = list(Chem.CanonicalRankAtoms(mol, breakTies=False))
    groups: dict[int, list[int]] = {}
    for atom in mol.GetAtoms():
        if atom.GetSymbol() == "C":
            groups.setdefault(ranks[atom.GetIdx()], []).append(atom.GetIdx())
    return list(groups.values())


def load() -> list[Delta50Compound]:
    """Every compound that passes the symmetry-class check."""
    raw = json.loads(DATA.read_text(encoding="utf-8"))
    out = []
    for name, rows in sorted(raw.items()):
        if name in EXCLUDED or name not in SMILES:
            continue
        mol = Chem.MolFromSmiles(SMILES[name])
        if mol is None or len(carbon_classes(mol)) != len(rows):
            continue
        out.append(
            Delta50Compound(
                name=name,
                smiles=SMILES[name],
                entries=[(str(a), float(b), float(c)) for a, b, c in rows],
            )
        )
    return out


@dataclass(frozen=True)
class Mapping:
    """An atom-index assignment plus how much to trust it."""

    shifts: dict[int, tuple[str, float]]
    #: Worst |our shielding - their shielding| after matching, in ppm. Two
    #: calculations at comparable levels should agree closely; a large value
    #: means the correspondence is not established.
    worst_shielding_gap: float
    #: Smallest gap between adjacent classes in OUR shieldings. When two
    #: classes are nearly degenerate the ordering could swap -- though so
    #: are their experimental shifts, which bounds the damage.
    closest_pair: float


def map_to_atoms(compound: Delta50Compound, our_shieldings: dict[int, float]) -> Mapping | None:
    """Assign DELTA50's rows to our atom indices via the two calculations.

    Both are isotropic shieldings of the same molecule at comparable
    levels, so ordering them and matching in order is a correspondence
    between structures, not between a prediction and an answer. That is
    what keeps the ground truth independent of the lookup.
    """
    mol = Chem.MolFromSmiles(compound.smiles)
    classes = carbon_classes(mol)
    if len(classes) != len(compound.entries):
        return None

    ours = []
    for group in classes:
        values = [our_shieldings[i] for i in group if i in our_shieldings]
        if not values:
            return None
        ours.append((sum(values) / len(values), group))
    ours.sort(key=lambda pair: pair[0])
    theirs = sorted(compound.entries, key=lambda e: e[2])

    shifts: dict[int, tuple[str, float]] = {}
    worst = 0.0
    for (our_shielding, group), (label, experimental, their_shielding) in zip(
        ours, theirs, strict=True
    ):
        worst = max(worst, abs(our_shielding - their_shielding))
        for index in group:
            shifts[index] = (label, experimental)
    gaps = [b[0] - a[0] for a, b in zip(ours, ours[1:], strict=False)]
    return Mapping(shifts=shifts, worst_shielding_gap=worst,
                   closest_pair=min(gaps) if gaps else float("inf"))
